from flask import Flask, render_template, request, redirect, url_for, flash
import json
from datetime import datetime, time
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = 'cheie_secreta_pentru_flash'

# 6 angajati, ca să corespundă cu QR-urile generate
angajati = {
    "emp1": {"nume": "Mihai Popa", "pin": "1111", "lucreaza_de_acasa": False},
    "emp2": {"nume": "Maria Ionescu", "pin": "2222", "lucreaza_de_acasa": True},
    "emp3": {"nume": "Ion Popescu", "pin": "3333", "lucreaza_de_acasa": False},
    "emp4": {"nume": "Ana Georgescu", "pin": "4444", "lucreaza_de_acasa": False},
    "emp5": {"nume": "Bogdan Marinescu", "pin": "5555", "lucreaza_de_acasa": True},
    "emp6": {"nume": "Claudia Radu", "pin": "6666", "lucreaza_de_acasa": False},
}

log_file = "angajati.json"

def log_event(qr_id, pin, act):
    if qr_id not in angajati:
        return "Cod QR invalid!"
    if angajati[qr_id]["pin"] != pin:
        return "PIN greșit!"

    azi = datetime.now().strftime("%Y-%m-%d")
    ora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if os.path.exists(log_file):
        with open(log_file, "r") as f:
            data = json.load(f)
    else:
        data = {}

    if azi not in data:
        data[azi] = {}

    if qr_id not in data[azi]:
        data[azi][qr_id] = {
            "nume": angajati[qr_id]["nume"],
            "check_in": None,
            "check_out": None,
            "pauza_masa": False,
            "de_acasa": angajati[qr_id]["lucreaza_de_acasa"]
        }

    if act == "check_in":
        data[azi][qr_id]["check_in"] = ora
    elif act == "check_out":
        data[azi][qr_id]["check_out"] = ora
    elif act == "pauza":
        now_time = datetime.now().time()
        if time(12, 0) <= now_time <= time(12, 30):
            data[azi][qr_id]["pauza_masa"] = True
        else:
            return "Pauza de masă poate fi luată doar între 12:00 și 12:30."

    with open(log_file, "w") as f:
        json.dump(data, f, indent=4)

    return None

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        act = request.form.get("act")
        return redirect(url_for("dashboard", act=act))
    return render_template("index.html")

@app.route("/dashboard/<act>", methods=["GET", "POST"])
def dashboard(act):
    if request.method == "POST":
        pin = request.form.get("pin")
        qr_id = request.form.get("qr")
        eroare = log_event(qr_id, pin, act)
        if eroare:
            flash(eroare)
            return redirect(url_for("dashboard", act=act))
        flash(f"{act.replace('_', ' ').title()} reușit pentru angajatul cu cod: {qr_id}")
        return redirect(url_for("index"))
    return render_template("dashboard.html", act=act)

@app.route("/admin")
def admin():
    if os.path.exists(log_file):
        with open(log_file, "r") as f:
            data = json.load(f)
    else:
        data = {}

    # Generare raport Excel detaliat
    wb = Workbook()
    wb.remove(wb.active)  # ștergem sheet-ul gol implicit

    for zi, angajati_zi in data.items():
        ws = wb.create_sheet(title=zi)

        # Antet tabel
        headers = ["Cod QR", "Nume", "Check-in", "Check-out", "Ore lucrate", "Întârziere", "Pauză masă", "Lucrează de acasă"]
        ws.append(headers)

        # Stilizare antet
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_num in range(1, len(headers)+1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        def parse_datetime(dt_str):
            try:
                return datetime.strptime(dt_str, '%Y-%m-%d %H:%M:%S')
            except Exception:
                return None

        # Populare date și calcule suplimentare
        for qr_id, info in angajati_zi.items():
            dt_in = parse_datetime(info.get("check_in")) if info.get("check_in") else None
            dt_out = parse_datetime(info.get("check_out")) if info.get("check_out") else None

            if dt_in and dt_out:
                delta = dt_out - dt_in
                ore_lucrate = round(delta.total_seconds() / 3600, 2)
            else:
                ore_lucrate = None

            intarziere = dt_in.time() > time(9, 0) if dt_in else False
            pauza_valida = info.get("pauza_masa", False)
            de_acasa = info.get("de_acasa", False)

            row = [
                qr_id,
                info.get("nume"),
                info.get("check_in"),
                info.get("check_out"),
                ore_lucrate,
                "Da" if intarziere else "Nu",
                "Da" if pauza_valida else "Nu",
                "Da" if de_acasa else "Nu"
            ]
            ws.append(row)

            # Actualizăm în dict pentru afișare HTML
            info["ore_lucrate"] = ore_lucrate
            info["intarziere"] = intarziere
            info["pauza_valida"] = pauza_valida

        # Ajustăm lățimea coloanelor
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    cell_value = str(cell.value)
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except Exception:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col_letter].width = adjusted_width

    raport_file = "raport_pontaj.xlsx"
    wb.save(raport_file)

    return render_template("admin.html", data=data, raport_file=raport_file)

if __name__ == "__main__":
    app.run(debug=True)
